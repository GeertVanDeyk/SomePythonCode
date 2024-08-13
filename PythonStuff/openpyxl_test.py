# -*- coding: utf-8 -*-
"""
Created on Wed Jun 26 08:09:45 2019

@author: gvandeyk
"""

import os
from openpyxl import Workbook, load_workbook

# Definition of directory structure
mijn_path = 'C:/Users/gvandeyk/Desktop/SKF Spec review/python_test'
mijn_input_path = mijn_path + '/' + 'input'
mijn_output_path = mijn_path + '/' + 'output'
mijn_output_filename = mijn_output_path + '/' + 'test.xls'

# Create output workbook
wb_out = Workbook()
sheet_out = wb_out.create_sheet(title="Summary")
wb_out.active = sheet_out
output_row = 1

# Loop accross all files in the input directory
# to collect data

for allFiles in os.listdir(mijn_input_path):
    mijn_input_workbook = allFiles
    wb_in = load_workbook(mijn_input_path + "/" + mijn_input_workbook)
    # display basic information -->
    print(os.path.basename(mijn_input_path + "/" + mijn_input_workbook))
    print('Available sheets : ' + str(wb_in.sheetnames))
    # <-- display basic information
    sheet = wb_in.active
    # display basic information -->
    print('Active sheet : ' + str(sheet))
    # <-- display basic information
    CellsContainingData = sheet.calculate_dimension()
    # display basic information -->
    print('Cells containing data : ' + CellsContainingData)
    print(CellsContainingData.split(':')[1])
    # <-- display basic information
    output_column = 1
# Loop over all cells in the input file containing data
# and write it to the output workbook
    for AllRows in sheet['A7' + ':' + CellsContainingData.split(':')[1]]:
        sheet_out.cell(output_row, output_column).value = sheet['B1'].value
        for AllCells in AllRows:
            output_column = output_column + 1
            sheet_out.cell(output_row, output_column).value = AllCells.value
        output_row = output_row + 1
        output_column = 1

wb_out.save(filename=mijn_output_filename)
